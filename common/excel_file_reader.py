
import openpyxl


def excel_file_read(file_name):
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    rows = []

    for row in range(2, max_row+1):
        tmp_row = []
        for col in range(1, max_column+1):
            cell_obj = sheet_obj.cell(row=row, column=col)
            tmp_row.append(cell_obj.value)
            # print(cell_obj.value)
        rows.append(tmp_row)

    return rows


def specific_columns_by_value_extract(data, column_name, value_check, extract_column):
    l_g_arr = []
    l_rs_arr = []
    print(column_name)
    col_index = extract_column.index(column_name)

    if isinstance(value_check, list):
        l_data = filter(extract_column, data)
        for l_row_1 in l_data.iterrows():
            l_g_arr = []
            if l_row_1[1][col_index] in value_check:
                for i in range(len(extract_column)):
                    l_g_arr.append(l_row_1[1][i])
                l_rs_arr.append(l_g_arr)
    return l_rs_arr


def headers_extract(file_name):
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    row = []

    for i in range(1, max_column + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        row.append(cell_obj)

    return row


if __name__ == "__main__":
    rows = excel_file_read(
        "../data/output/bom/List_of_downloaded_crs_2022-05-25.xlsx")

    for i in range(len(rows)):
        print(rows[i])
